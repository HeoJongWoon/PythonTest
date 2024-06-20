import os
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side

def createFolder():
    # 현재 날짜를 YYYY-MM 형식으로 변환
    currentDate = datetime.now().strftime("%Y-%m")

    # 기본 폴더 경로 설정
    basePath = "C:/RPA/지자체 희망일자리"
    folderPath = os.path.join(basePath, currentDate)

    # RPA 폴더가 존재하지 않으면 오류 반환
    if not os.path.exists("C:/RPA"):
        # RPA 폴더가 없음을 알리는 메시지 설정
        msg = "RPA 폴더가 존재하지 않습니다."
        print(msg)
        return msg

    # 지자체 희망일자리 폴더가 존재하지 않으면 오류 반환
    if not os.path.exists(basePath):
        # 지자체 희망일자리 폴더가 없음을 알리는 메시지 설정
        msg = "지자체 희망일자리 폴더가 존재하지 않습니다."
        print(msg)
        return msg

    # 날짜 폴더가 존재하지 않으면 생성
    if not os.path.exists(folderPath):
        # 날짜 폴더 생성
        os.makedirs(folderPath)
        msg = "성공"
        print(msg)
    else:
        # 날짜 폴더가 이미 존재하면 성공 메시지 반환
        msg = "성공"
        print(msg)

    return msg

def createFile():
    # 현재 날짜를 YYYY-MM 형식으로 변환
    currentDate = datetime.now().strftime("%Y-%m")

    # 기본 폴더 경로 설정
    basePath = "C:/RPA/지자체 희망일자리"
    folderPath = os.path.join(basePath, currentDate)
    filePath = os.path.join(folderPath, f"{currentDate}.xlsx")

    # RPA 폴더가 존재하지 않으면 오류 반환
    if not os.path.exists("C:/RPA"):
        # RPA 폴더가 없음을 알리는 메시지 설정
        msg = "RPA 폴더가 존재하지 않습니다."
        print(msg)
        return msg

    # 지자체 희망일자리 폴더가 존재하지 않으면 오류 반환
    if not os.path.exists(basePath):
        # 지자체 희망일자리 폴더가 없음을 알리는 메시지 설정
        msg = "지자체 희망일자리 폴더가 존재하지 않습니다."
        print(msg)
        return msg

    # 날짜 폴더가 존재하지 않으면 오류 반환
    if not os.path.exists(folderPath):
        # 날짜 폴더가 없음을 알리는 메시지 설정
        msg = f"{folderPath} 폴더가 존재하지 않습니다."
        print(msg)
        return msg

    # 엑셀 파일이 존재하지 않으면 생성
    if not os.path.exists(filePath):
        # 새로운 엑셀 워크북 생성
        workbook = openpyxl.Workbook()
        # 기본 활성화된 시트 선택
        sheet = workbook.active

        # 컬럼 헤더 설정
        headers = ["구분", "사업명", "신청기간", "근무지", "임금조건(보수)", "URL", "등록일", "문의전화"]
        # 시트에 헤더 추가
        sheet.append(headers)

        # 스타일 설정
        # 셀 배경색 설정
        fill = PatternFill(start_color="FAC090", end_color="FAC090", fill_type="solid")
        # 글자 볼드체 설정
        font = Font(bold=True)
        # 셀 테두리 설정 (좌, 우, 상, 하)
        border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))

        # 헤더 셀에 스타일 적용
        for cell in sheet[1]:
            cell.fill = fill  # 배경색 적용
            cell.font = font  # 볼드체 적용
            cell.border = border  # 테두리 적용

        # 엑셀 파일 저장
        workbook.save(filePath)
        msg = "성공"
        print(msg)
    else:
        # 엑셀 파일이 이미 존재하면 성공 메시지 반환
        msg = "성공"
        print(msg)

    return msg

# 폴더와 파일 생성 예제 실행
createFolder()
createFile()
